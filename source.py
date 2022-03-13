"""
RNS Generator v1.0.1
    A Python script for generating random number sequences from an external reference table. This script can be used for random, systematic and monetary 
unit sampling.
    For installation, use and troubleshooting, please refer to the README file. Code is available in the author's public Github repository by following the 
link below.

(C) 2021 Carmelo Julius Paz
(email) carmelopaz0708@gmail.com
(github) https://github.com/carmelopaz0708

Released under the GNU General Public License (gpl-3.0)
"""

from openpyxl import load_workbook
import os, csv, sys

def main():
    # Ensure proper usage
    if len(sys.argv) != 2:
        print("Usage: python source.py reference")
        sys.exit(1)
    
    # Check correct file extension
    allowableFileTypes = ('.xlsx', '.xlsm', '.xltx', '.xltm')
    if not sys.argv[1].endswith(allowableFileTypes):
        print("Invalid file type used as reference.")
        sys.exit(2)

    # Check if tables directory exists
    referenceDir = "tables"
    if not os.path.isdir(referenceDir):
        os.makedirs(referenceDir)
        print("Generated table directory in root folder. Please save your tables here.")

    # Check if file is in tables directory
    referenceFile = os.path.join(referenceDir, sys.argv[1])
    if not os.path.exists(referenceFile):
        print("Reference file \"{}\" missing from tables directory.".format(sys.argv[1]))
        sys.exit(3)

    sep = "----------------------------------------------------------"
    print("{}\n\tRANDOM NUMBER SEQUENCE GENERATOR\n{}".format(sep, sep))

    # Load variables into memory.
    wb = load_workbook(referenceFile)
    
    tbName = getTableName(wb, sep)
    tbData = getTableData(wb, tbName)
    pSize = getPopulation(sep)
    sSize = getSampleSize(sep)
    rowLength = len(tbData)
    colLength = len(tbData[0])
    pos = getPosition(rowLength, colLength, tbData, sep)
    swp = getSweep(sep)

    # Check user input
    isFinal = finalize(tbName, pSize, sSize, pos, swp, sep)
    
    if isFinal == False:
        print("Closing program.")
        sys.exit(0)

    outSeq = []

    # Generate RNS sequence with Left to Right sweep
    if swp == "Left to Right":
        slicedTab = tbData[pos[1]:]
        slicedTab[0] = slicedTab[0][pos[2]:]
        
        for row in slicedTab:
            for cell in row:
                if cell == 0 or cell in outSeq:
                    continue

                if cell <= pSize and len(outSeq) < sSize:
                    outSeq.append(cell)

        if len(outSeq) < sSize:
            print("Temp: {}".format(outSeq))
            isLoopBack = promptLoopBack(sep)

            if isLoopBack:
                for row in tbData:
                    for cell in row:
                        if cell == 0 or cell in outSeq:
                            continue

                        if cell <= pSize and len(outSeq) < sSize:
                            outSeq.append(cell)

    # Generate RNS sequence with Down, Left to Right sweep
    if swp == "Down, Left to Right":
        zipTab = [*zip(*tbData)]
        slicedTab = zipTab[pos[2]:]
        slicedTab[0] = slicedTab[0][pos[1]:]

        for row in slicedTab:
            for cell in row:
                if cell == 0 or cell in outSeq:
                    continue

                if cell <= pSize and len(outSeq) < sSize:
                    outSeq.append(cell)

        if len(outSeq) < sSize:
            print("Temp: {}".format(outSeq))
            isLoopBack = promptLoopBack(sep)

            if isLoopBack:
                for row in zipTab:
                    for cell in row:
                        if cell == 0 or cell in outSeq:
                            continue

                        if cell <= pSize and len(outSeq) < sSize:
                            outSeq.append(cell)

    print("Output: {}".format(outSeq))

    # Export as .txt
    outDir = "output"
    if not os.path.exists(outDir):
        os.makedirs(outDir)

    outFile = "out.txt"
    outPath = os.path.join(outDir, outFile)

    with open(outPath, "w") as file:
        file.write("{}\nWorkbook: {}".format(sep, sys.argv[1]))
        file.write("\nWorksheet: {}".format(tbName))
        file.write("\nPopulation: {}".format(pSize))
        file.write("\nSample: {}".format(sSize))
        file.write("\nRNS Start Value: {}".format(pos[0]))
        file.write("\nRNS Start Position: row {}, col. {}".format(pos[1] + 1, pos[2] + 1))
        file.write("\nSweep: {}".format(swp))
        file.write("\n{}\nOUTPUT:\n".format(sep))
        wr = csv.writer(file)
        wr.writerow(outSeq)
    
    print("Exit with success. Please check {} in the {} folder of root directory.".format(outFile, outDir))
    sys.exit(0)


def getTableName(workbook, s):
    '''Returns the name of the worksheet'''

    sheets = tuple(workbook.sheetnames)

    print("INFO: Loaded {} sheets from {}\n{}\n\t\tENTER VARIABLES\nTable Names:".format(len(sheets), sys.argv[1], s))
    
    for index, title in enumerate(sheets):
        print("({}) {}".format(index + 1, title))

    while True:
        try:
            usrTable = int(input("Select desired table: "))

            if (usrTable - 1) < 0 or (usrTable - 1) > len(sheets):
                raise ValueError

            break

        except ValueError:
            print("Enter number from available options.")

        except KeyboardInterrupt:
            sys.exit(0)
    
    print(s)
    table = sheets[usrTable - 1]

    return table


def getTableData(key, value):
    '''Returns the cell values of worksheet_name as a 2D list'''

    ws = key[value]
    data = []

    for row in ws.iter_rows(min_row = 1, values_only = True):
        data.append(list(row))

    return data


def getPopulation(s):
    '''Returns user input for population size'''

    while True:
        try:
            populationLimit = int(input("Enter population limit(inclusive): "))

            if populationLimit <= 0:
                raise ValueError

            break

        except ValueError:
            print("Value must be a valid positive integer")

        except KeyboardInterrupt:
            sys.exit(0)

    print(s)

    return populationLimit


def getSampleSize(s):
    '''Returns user input for sample size'''

    while True:
        try:
            sampleSize = int(input("Enter sample size: "))

            if sampleSize <= 0:
                raise ValueError

            break

        except ValueError:
            print("Value must be a valid positive integer.")

        except KeyboardInterrupt:
            sys.exit(0)

    print(s)

    return sampleSize


def getPosition(maxRows, maxColumns, table, s):
    '''Obtains starting cell position from worksheet_name'''

    while True:
        try:
            usrRow = int(input("Enter starting row(-y): "))

            if usrRow <= 0 or usrRow > maxRows:
                raise ValueError

            break

        except ValueError:
            print("Value exceeds maximum number of rows in table.")

        except KeyboardInterrupt:
            sys.exit(0)

    while True:
        try:
            usrCol = int(input("Enter starting column(+x): "))

            if usrCol <= 0 or usrCol > maxColumns:
                raise ValueError

            break

        except ValueError:
            print("Value exceeds maximum number of columns in table.")

        except KeyboardInterrupt:
            sys.exit(0)

    print(s)
    rStart = table[usrRow - 1][usrCol - 1]

    return rStart, usrRow - 1, usrCol - 1


def getSweep(s):
    '''Returns general sweeping algorithm as string'''

    allowableSweep = ("Left to Right", "Down, Left to Right")
    print("Sweep patterns:")

    for index, sweep in enumerate(allowableSweep):
        print("({}) {}".format(index + 1, sweep))

    while True:
        try:
            usrSweep = int(input("Select desired sweep pattern: "))

            if (usrSweep - 1) < 0 or (usrSweep - 1) > len(allowableSweep):
                raise ValueError

            break

        except ValueError:
            print("Invalid sweep pattern option")

        except KeyboardInterrupt:
            sys.exit(0)

    print(s)

    return allowableSweep[usrSweep - 1]


def finalize(tableName, populationSize, sampleSize, position, sweep, s):
    '''Returns flag for proceed with sequence'''

    print("Table: {}\nPopulation Size: {}\nSample Size: {}".format(tableName, populationSize, sampleSize))
    print("RNS Start Value: {}\nRNS Start Position: row {}, col. {}".format(position[0], position[1] + 1, position[2] + 1))
    print("Sweep: {}".format(sweep))

    while True:
        try:
            response = input("Is this correct(y/n)? ").rstrip().lower()
            
            if response not in ['y', 'n']:
                continue

            else:
                print(s)

                if response == 'y':
                    return True

                else:
                    return False
        
        except KeyboardInterrupt:
            sys.exit(0)


def promptLoopBack(s):
    while True:
        try:
            isloopBack = input("Sequence less than sample size. Continue from start(y, n)? ").rstrip().lower()
            print(s)

            if isloopBack not in ['y', 'n']:
                continue

            else:

                if isloopBack == 'y':
                    return True

                else:
                    return False
        
        except KeyboardInterrupt:
            sys.exit(0)


if __name__ == "__main__":
    main()
